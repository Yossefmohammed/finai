"""
services/report_service.py
===========================
Generates professional PDF and Excel financial reports.
Called by /api/export/pdf and /api/export/excel endpoints.
"""

import io
from datetime import datetime
from sqlalchemy.orm import Session
from app.models.database import Transaction

# ── Shared data fetcher ────────────────────────────────────────────────────────
def _get_report_data(db: Session, user_id: int, period: str = "all") -> dict:
    """Fetch all data needed for a report."""
    from app.tools.financial_tools import (
        get_expenses, get_dashboard_summary,
        detect_anomalies, _parse_period
    )

    summary   = get_dashboard_summary(db, period, user_id=user_id)
    expenses  = get_expenses(db, period=period, type="expense", user_id=user_id)
    income    = get_expenses(db, period=period, type="income",  user_id=user_id)
    anomalies = detect_anomalies(db, user_id=user_id)

    # Raw transactions for table
    q = db.query(Transaction).filter(Transaction.user_id == user_id)
    start, end = _parse_period(period)
    if start:
        q = q.filter(Transaction.date >= start, Transaction.date <= end)
    transactions = q.order_by(Transaction.date.desc()).limit(50).all()

    return {
        "summary":      summary,
        "expenses":     expenses,
        "income":       income,
        "anomalies":    anomalies,
        "transactions": transactions,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "period":       period,
    }


# ══════════════════════════════════════════════════════════════════════════════
# PDF REPORT
# ══════════════════════════════════════════════════════════════════════════════
def generate_pdf(db: Session, user_id: int, business_name: str,
                 period: str = "all") -> bytes:
    """Generate a professional PDF financial report. Returns raw bytes."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    data   = _get_report_data(db, user_id, period)
    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                                rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    story  = []

    # ── Color palette ──────────────────────────────────────────────────────────
    BLUE   = colors.HexColor("#3b7bff")
    DARK   = colors.HexColor("#111318")
    GRAY   = colors.HexColor("#f4f5f7")
    GREEN  = colors.HexColor("#22c55e")
    RED    = colors.HexColor("#ef4444")
    WHITE  = colors.white

    # ── Custom styles ──────────────────────────────────────────────────────────
    title_style = ParagraphStyle("Title", parent=styles["Title"],
                                  textColor=DARK, fontSize=22, spaceAfter=4)
    sub_style   = ParagraphStyle("Sub", parent=styles["Normal"],
                                  textColor=colors.HexColor("#6b7280"),
                                  fontSize=10, spaceAfter=2)
    h2_style    = ParagraphStyle("H2", parent=styles["Heading2"],
                                  textColor=BLUE, fontSize=13, spaceBefore=14, spaceAfter=6)
    normal      = ParagraphStyle("Body", parent=styles["Normal"],
                                  fontSize=9, leading=14)

    fmt = lambda n: f"{n:,.2f} EGP"

    # ── Header ─────────────────────────────────────────────────────────────────
    story.append(Paragraph("FinAI", ParagraphStyle("Logo", parent=styles["Title"],
                 textColor=BLUE, fontSize=28, spaceAfter=2)))
    story.append(Paragraph("AI Financial Report", title_style))
    story.append(Paragraph(f"{business_name}  |  Period: {period}  |  Generated: {data['generated_at']}", sub_style))
    story.append(HRFlowable(width="100%", thickness=2, color=BLUE, spaceAfter=12))

    # ── KPI summary cards (table layout) ──────────────────────────────────────
    story.append(Paragraph("Financial Summary", h2_style))
    s = data["summary"]
    profit_color = GREEN if s["profit"] >= 0 else RED

    kpi_data = [
        ["Total Income", "Total Expenses", "Net Profit", "Transactions"],
        [fmt(s["income"]), fmt(s["expenses"]), fmt(s["profit"]), str(s["transaction_count"])],
    ]
    kpi_table = Table(kpi_data, colWidths=[4.2*cm]*4)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), BLUE),
        ("TEXTCOLOR",    (0,0), (-1,0), WHITE),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,0), 9),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("BACKGROUND",   (0,1), (-1,1), GRAY),
        ("FONTNAME",     (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE",     (0,1), (-1,1), 11),
        ("TEXTCOLOR",    (2,1), (2,1), profit_color),
        ("ROWBACKGROUNDS",(0,1),(-1,1),[GRAY]),
        ("BOX",          (0,0), (-1,-1), 0.5, colors.HexColor("#d1d5db")),
        ("INNERGRID",    (0,0), (-1,-1), 0.5, colors.HexColor("#d1d5db")),
        ("TOPPADDING",   (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0), (-1,-1), 8),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 16))

    # ── Top categories ─────────────────────────────────────────────────────────
    story.append(Paragraph("Top Spending Categories", h2_style))
    cats = s.get("top_categories", [])
    if cats:
        cat_data = [["#", "Category", "Amount"]]
        for i, c in enumerate(cats, 1):
            cat_data.append([str(i), c["category"], fmt(c["amount"])])
        cat_table = Table(cat_data, colWidths=[1*cm, 10*cm, 5*cm])
        cat_table.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), DARK),
            ("TEXTCOLOR",    (0,0), (-1,0), WHITE),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 9),
            ("ALIGN",        (0,0), (0,-1), "CENTER"),
            ("ALIGN",        (2,0), (2,-1), "RIGHT"),
            ("ROWBACKGROUNDS",(1,0),(-1,-1),[WHITE, GRAY]),
            ("BOX",          (0,0), (-1,-1), 0.5, colors.HexColor("#d1d5db")),
            ("INNERGRID",    (0,0), (-1,-1), 0.3, colors.HexColor("#e5e7eb")),
            ("TOPPADDING",   (0,0), (-1,-1), 6),
            ("BOTTOMPADDING",(0,0), (-1,-1), 6),
            ("LEFTPADDING",  (0,0), (-1,-1), 8),
        ]))
        story.append(cat_table)
    else:
        story.append(Paragraph("No category data available.", normal))
    story.append(Spacer(1, 16))

    # ── Anomalies ──────────────────────────────────────────────────────────────
    anomalies = data["anomalies"].get("anomalies", [])
    if anomalies:
        story.append(Paragraph(f"Anomalies Detected ({len(anomalies)})", h2_style))
        an_data = [["Date", "Description", "Amount", "Z-Score"]]
        for a in anomalies[:10]:
            an_data.append([a["date"], a["description"][:40],
                             fmt(a["amount"]), str(a["z_score"])])
        an_table = Table(an_data, colWidths=[2.5*cm, 8*cm, 4*cm, 2*cm])
        an_table.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), RED),
            ("TEXTCOLOR",    (0,0), (-1,0), WHITE),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(1,0),(-1,-1),[colors.HexColor("#fff5f5"), WHITE]),
            ("BOX",          (0,0), (-1,-1), 0.5, RED),
            ("INNERGRID",    (0,0), (-1,-1), 0.3, colors.HexColor("#fecaca")),
            ("TOPPADDING",   (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ("LEFTPADDING",  (0,0), (-1,-1), 6),
        ]))
        story.append(an_table)
        story.append(Spacer(1, 16))

    # ── Recent transactions ────────────────────────────────────────────────────
    story.append(Paragraph("Recent Transactions (last 50)", h2_style))
    txs = data["transactions"]
    if txs:
        tx_data = [["Date", "Description", "Category", "Type", "Amount"]]
        for tx in txs:
            tx_data.append([
                tx.date.strftime("%Y-%m-%d"),
                tx.description[:35] if tx.description else "",
                tx.category or "",
                tx.type or "",
                fmt(tx.amount),
            ])
        tx_table = Table(tx_data, colWidths=[2.3*cm, 6.5*cm, 3.5*cm, 2*cm, 3.2*cm])
        tx_table.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), DARK),
            ("TEXTCOLOR",    (0,0), (-1,0), WHITE),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 8),
            ("ALIGN",        (4,0), (4,-1), "RIGHT"),
            ("ROWBACKGROUNDS",(1,0),(-1,-1),[WHITE, GRAY]),
            ("BOX",          (0,0), (-1,-1), 0.5, colors.HexColor("#d1d5db")),
            ("INNERGRID",    (0,0), (-1,-1), 0.3, colors.HexColor("#e5e7eb")),
            ("TOPPADDING",   (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ("LEFTPADDING",  (0,0), (-1,-1), 6),
        ]))
        story.append(tx_table)
    else:
        story.append(Paragraph("No transactions found.", normal))

    # ── Footer ─────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=1, color=GRAY))
    story.append(Paragraph(
        f"Generated by FinAI  •  {data['generated_at']}  •  Confidential",
        ParagraphStyle("Footer", parent=styles["Normal"],
                       textColor=colors.HexColor("#9ca3af"),
                       fontSize=8, alignment=TA_CENTER, spaceBefore=6)
    ))

    doc.build(story)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════════════
def generate_excel(db: Session, user_id: int, business_name: str,
                   period: str = "all") -> bytes:
    """Generate a professional Excel workbook. Returns raw bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint

    data = _get_report_data(db, user_id, period)
    s    = data["summary"]
    wb   = Workbook()

    # ── Shared styles ──────────────────────────────────────────────────────────
    BLUE_FILL  = PatternFill("solid", fgColor="3B7BFF")
    DARK_FILL  = PatternFill("solid", fgColor="111318")
    GRAY_FILL  = PatternFill("solid", fgColor="F4F5F7")
    GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")
    RED_FILL   = PatternFill("solid", fgColor="FEE2E2")
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    BOLD       = Font(bold=True)
    HEADER_BORDER = Border(
        bottom=Side(style="medium", color="3B7BFF")
    )
    MONEY_FMT  = '#,##0.00 "EGP"'
    CENTER     = Alignment(horizontal="center", vertical="center")
    RIGHT_AL   = Alignment(horizontal="right")

    def hdr(ws, row, col, val, fill=DARK_FILL, font=WHITE_FONT):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill; c.font = font; c.alignment = CENTER
        return c

    def money(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = MONEY_FMT; c.alignment = RIGHT_AL
        return c

    # ══ Sheet 1: Summary ═══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 22

    # Title block
    ws1.merge_cells("A1:B1")
    c = ws1["A1"]; c.value = "FinAI Financial Report"
    c.font = Font(bold=True, size=18, color="3B7BFF"); c.alignment = CENTER
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:B2")
    c = ws1["A2"]; c.value = f"{business_name}  |  {period}  |  {data['generated_at']}"
    c.font = Font(size=10, color="6B7280"); c.alignment = CENTER
    ws1.row_dimensions[2].height = 20

    # KPI section
    kpis = [
        ("Total Income",      s["income"],    GREEN_FILL, "22C55E"),
        ("Total Expenses",    s["expenses"],  RED_FILL,   "EF4444"),
        ("Net Profit",        s["profit"],    GRAY_FILL,  "3B7BFF"),
        ("Total Transactions",s["transaction_count"], GRAY_FILL, "111318"),
    ]
    for i, (label, val, fill, color) in enumerate(kpis, start=4):
        ws1.row_dimensions[i].height = 26
        lc = ws1.cell(row=i, column=1, value=label)
        lc.font = Font(bold=True, size=11)
        lc.fill = fill; lc.alignment = Alignment(horizontal="left", indent=1)

        vc = ws1.cell(row=i, column=2, value=val)
        vc.font = Font(bold=True, size=12, color=color)
        vc.fill = fill
        if i < 7:
            vc.number_format = MONEY_FMT
        vc.alignment = RIGHT_AL

    # Top categories table
    row = 10
    ws1.cell(row=row, column=1, value="Top Categories").font = Font(bold=True, size=12, color="3B7BFF")
    row += 1
    hdr(ws1, row, 1, "Category"); hdr(ws1, row, 2, "Amount")
    for cat in s.get("top_categories", []):
        row += 1
        ws1.cell(row=row, column=1, value=cat["category"])
        money(ws1, row, 2, cat["amount"])
        if row % 2 == 0:
            for col in [1, 2]:
                ws1.cell(row=row, column=col).fill = GRAY_FILL

    # ── Bar chart ──────────────────────────────────────────────────────────────
    cats = s.get("top_categories", [])
    if cats:
        chart_row = 10
        chart_start = 10 + 1 + len(cats) + 1  # after the table
        chart = BarChart()
        chart.type = "col"
        chart.title = "Spending by Category"
        chart.y_axis.title = "Amount (EGP)"
        chart.x_axis.title = "Category"
        chart.height = 12; chart.width = 18
        chart.style = 10

        data_ref = Reference(ws1, min_col=2, min_row=11, max_row=11+len(cats))
        cats_ref  = Reference(ws1, min_col=1, min_row=12, max_row=11+len(cats))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws1.add_chart(chart, f"D4")

    # ══ Sheet 2: Transactions ══════════════════════════════════════════════════
    ws2 = wb.create_sheet("Transactions")
    ws2.sheet_view.showGridLines = False
    headers = ["Date", "Description", "Category", "Type", "Amount", "Currency"]
    widths  = [14, 40, 20, 12, 18, 10]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        hdr(ws2, 1, col, h)
        ws2.column_dimensions[get_column_letter(col)].width = w

    for i, tx in enumerate(data["transactions"], start=2):
        ws2.cell(row=i, column=1, value=tx.date.strftime("%Y-%m-%d")).alignment = CENTER
        ws2.cell(row=i, column=2, value=tx.description or "")
        ws2.cell(row=i, column=3, value=tx.category or "")
        ws2.cell(row=i, column=4, value=tx.type or "").alignment = CENTER
        money(ws2, i, 5, tx.amount)
        ws2.cell(row=i, column=6, value=tx.currency or "EGP").alignment = CENTER
        if i % 2 == 0:
            for col in range(1, 7):
                ws2.cell(row=i, column=col).fill = GRAY_FILL
        if tx.type == "income":
            ws2.cell(row=i, column=4).fill = PatternFill("solid", fgColor="DCFCE7")
        elif tx.type == "expense":
            ws2.cell(row=i, column=4).fill = PatternFill("solid", fgColor="FEE2E2")

    # ══ Sheet 3: Anomalies ═════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Anomalies")
    ws3.sheet_view.showGridLines = False
    an_headers = ["Date", "Description", "Category", "Amount", "Z-Score"]
    an_widths  = [14, 40, 20, 18, 10]
    for col, (h, w) in enumerate(zip(an_headers, an_widths), 1):
        hdr(ws3, 1, col, h, fill=PatternFill("solid", fgColor="EF4444"))
        ws3.column_dimensions[get_column_letter(col)].width = w

    anomalies = data["anomalies"].get("anomalies", [])
    if anomalies:
        for i, a in enumerate(anomalies, start=2):
            ws3.cell(row=i, column=1, value=a["date"]).alignment = CENTER
            ws3.cell(row=i, column=2, value=a["description"])
            ws3.cell(row=i, column=3, value=a["category"])
            money(ws3, i, 4, a["amount"])
            ws3.cell(row=i, column=5, value=a["z_score"]).alignment = CENTER
            for col in range(1, 6):
                ws3.cell(row=i, column=col).fill = PatternFill("solid", fgColor="FFF5F5")
    else:
        ws3.cell(row=2, column=1, value="No anomalies detected.").font = Font(color="22C55E", bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()